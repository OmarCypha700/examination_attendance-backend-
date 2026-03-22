import csv
import io

import openpyxl
from django.contrib.auth import get_user_model
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import filters, generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import CanScan, IsAdmin, IsAdminOrReadOnly

from .models import Course, ExamAttendance, ExamSession, Level, Program, Student
from .serializers import (
    AttendanceSerializer,
    CourseSerializer,
    ExamSessionSerializer,
    LevelSerializer,
    ProgramSerializer,
    ScanSerializer,
    StudentSerializer,
)

User = get_user_model()

# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
CENTER = Alignment(horizontal="center")


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _parse_uploaded_file(file):
    """Return list of dicts from an uploaded CSV or XLSX file."""
    name = file.name.lower()
    if name.endswith(".csv"):
        decoded = file.read().decode("utf-8-sig")  # handle BOM
        reader = csv.DictReader(io.StringIO(decoded))
        return [row for row in reader]
    elif name.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append(
                {
                    headers[i]: (str(v).strip() if v is not None else "")
                    for i, v in enumerate(row)
                }
            )
        return result
    else:
        raise ValueError("Unsupported file format. Upload a CSV or XLSX file.")


# ─────────────────────────────────────────────────────────────────────────────
# Programs & Levels
# ─────────────────────────────────────────────────────────────────────────────


class ProgramListCreateView(generics.ListCreateAPIView):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer
    permission_classes = (IsAdminOrReadOnly,)
    filter_backends = (filters.SearchFilter,)
    search_fields = ("name", "code")
    pagination_class = None


class ProgramDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer
    permission_classes = (IsAdminOrReadOnly,)


class LevelListCreateView(generics.ListCreateAPIView):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer
    permission_classes = (IsAdminOrReadOnly,)
    pagination_class = None


class LevelDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer
    permission_classes = (IsAdminOrReadOnly,)


# ─────────────────────────────────────────────────────────────────────────────
# Students
# ─────────────────────────────────────────────────────────────────────────────


class StudentListCreateView(generics.ListCreateAPIView):
    serializer_class = StudentSerializer
    permission_classes = (IsAdminOrReadOnly,)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter)
    filterset_fields = ("programme", "level", "is_active")
    search_fields = ("index_number", "full_name")

    def get_queryset(self):
        return Student.objects.select_related("programme", "level").all()


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StudentSerializer
    permission_classes = (IsAdminOrReadOnly,)

    def get_queryset(self):
        return Student.objects.select_related("programme", "level").all()


class StudentBulkCreateView(APIView):
    permission_classes = (IsAdmin,)

    def post(self, request):
        items = request.data
        if not isinstance(items, list):
            return Response({"detail": "Expected a list."}, status=400)

        created_count, errors = 0, []
        students_to_create = []

        for i, item in enumerate(items):
            serializer = StudentSerializer(data=item)
            if serializer.is_valid():
                students_to_create.append(Student(**serializer.validated_data))
            else:
                errors.append({"row": i + 1, "errors": serializer.errors})

        if students_to_create:
            created_students = Student.objects.bulk_create(
                students_to_create, ignore_conflicts=True
            )
            created_count = len(created_students)

        return Response(
            {"created": created_count, "errors": errors},
            status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED,
        )


# ── Student Export ────────────────────────────────────────────────────────────


class StudentExportView(APIView):
    """
    GET /api/students/export/?format=csv|xlsx&template=true
    template=true  → headers-only file (blank import template)
    """

    permission_classes = (IsAdmin,)

    HEADERS = ["index_number", "full_name", "programme_code", "level_name"]

    def get(self, request):
        fmt = request.query_params.get("format", "xlsx").lower()
        template_only = request.query_params.get("template") == "true"
        qs = (
            []
            if template_only
            else Student.objects.select_related("programme", "level").all()
        )

        if fmt == "csv":
            return self._csv(qs, template_only)
        return self._xlsx(qs, template_only)

    def _xlsx(self, qs, template_only):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(self.HEADERS)
        _style_header_row(ws, 1, len(self.HEADERS))
        for col, width in zip("ABCDE", [20, 35, 18, 15, 8]):
            ws.column_dimensions[col].width = width
        if not template_only:
            for s in qs:
                ws.append(
                    [
                        s.index_number,
                        s.full_name,
                        s.programme.code,
                        s.level.name,
                        # s.gender,
                    ]
                )
        name = "students_template.xlsx" if template_only else "students.xlsx"
        return _xlsx_response(wb, name)


# ── Student Import ────────────────────────────────────────────────────────────


class StudentImportView(APIView):
    """
    POST /api/students/import/
    Multipart file upload. Columns: index_number, full_name, programme_code, level_name
    Rows are upserted by index_number.
    """

    permission_classes = (IsAdmin,)
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            rows = _parse_uploaded_file(file)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        if not rows:
            return Response({"detail": "The file is empty."}, status=400)

        # pre-fetch lookup maps
        programmes = {p.code.upper(): p for p in Program.objects.all()}
        levels = {level.name.upper(): level for level in Level.objects.all()}

        created, updated, errors = 0, 0, []

        for i, row in enumerate(rows, start=1):
            index_number = row.get("index_number", "").strip()
            full_name = row.get("full_name", "").strip()
            programme_code = row.get("programme_code", "").strip().upper()
            level_name = row.get("level_name", "").strip().upper()

            if not index_number:
                errors.append({"row": i, "error": "index_number is required."})
                continue
            if not full_name:
                errors.append({"row": i, "error": "full_name is required."})
                continue
            if programme_code not in programmes:
                errors.append(
                    {"row": i, "error": f"Programme code '{programme_code}' not found."}
                )
                continue
            if level_name not in levels:
                errors.append({"row": i, "error": f"Level '{level_name}' not found."})
                continue

            try:
                _, was_created = Student.objects.update_or_create(
                    index_number=index_number,
                    defaults={
                        "full_name": full_name,
                        "programme": programmes[programme_code],
                        "level": levels[level_name],
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        return Response(
            {"created": created, "updated": updated, "errors": errors},
            status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_200_OK,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Courses
# ─────────────────────────────────────────────────────────────────────────────


class CourseListCreateView(generics.ListCreateAPIView):
    serializer_class = CourseSerializer
    permission_classes = (IsAdminOrReadOnly,)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter)
    filterset_fields = ("programme", "level")
    search_fields = ("course_code", "course_title")

    def get_queryset(self):
        return Course.objects.select_related("programme", "level").all()


class CourseDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CourseSerializer
    permission_classes = (IsAdminOrReadOnly,)

    def get_queryset(self):
        return Course.objects.select_related("programme", "level").all()


class CourseExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses"

        # Header
        ws.append(
            [
                "course_code",
                "course_title",
                "programme_code",
                "level_name",
            ]
        )

        courses = Course.objects.select_related("programme", "level")

        for c in courses:
            ws.append(
                [
                    c.course_code,
                    c.course_title,
                    c.programme.code,
                    c.level.name,
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = "attachment; filename=courses.xlsx"

        wb.save(response)
        return response


class CourseImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):

        file = request.FILES.get("file")

        if not file:
            return Response({"error": "No file uploaded"}, status=400)

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        created = 0
        updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            course_code, course_title, programme_code, level_name = row

            if not course_code:
                continue

            try:
                programme = Program.objects.get(code=programme_code)
                level = Level.objects.get(name=level_name)
            except (Program.DoesNotExist, Level.DoesNotExist):
                continue

            obj, is_created = Course.objects.update_or_create(
                course_code=course_code,
                defaults={
                    "course_title": course_title,
                    "programme": programme,
                    "level": level,
                },
            )

            if is_created:
                created += 1
            else:
                updated += 1

        return Response({"created": created, "updated": updated})


class CourseTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["course_code", "course_title", "programme_code", "level_name"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = "attachment; filename=courses_template.xlsx"

        wb.save(response)
        return response


# ─────────────────────────────────────────────────────────────────────────────
# Exam Sessions
# ─────────────────────────────────────────────────────────────────────────────


class ExamSessionListCreateView(generics.ListCreateAPIView):
    serializer_class = ExamSessionSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    )
    filterset_fields = ("status", "programme", "level", "date")
    search_fields = ("course__course_code", "course__course_title", "venue")
    ordering_fields = ("date", "start_time", "created_at")

    def get_queryset(self):
        return (
            ExamSession.objects.select_related(
                "course", "programme", "level", "created_by"
            )
            .annotate(attendance_count=Count("attendances"))
            .all()
        )

    def get_permissions(self):
        return [IsAdmin()] if self.request.method == "POST" else [IsAuthenticated()]


class ExamSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ExamSessionSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return ExamSession.objects.select_related(
            "course", "programme", "level", "created_by"
        ).all()

    def get_permissions(self):
        if self.request.method in ("PUT", "PATCH", "DELETE"):
            return [IsAdmin()]
        return [IsAuthenticated()]


class ExamSessionStatusView(APIView):
    """PATCH /api/exam-sessions/<pk>/status/"""

    permission_classes = (IsAdmin,)

    def patch(self, request, pk):
        session = get_object_or_404(ExamSession, pk=pk)
        new_status = request.data.get("status")

        if new_status not in ("scheduled", "active", "closed"):
            return Response({"detail": "Invalid status."}, status=400)

        if new_status == "active":
            conflict = (
                ExamSession.objects.filter(
                    programme=session.programme,
                    level=session.level,
                    status="active",
                )
                .exclude(pk=pk)
                .first()
            )
            if conflict:
                return Response(
                    {
                        "detail": f"Session '{conflict}' is already active for this programme/level."
                    },
                    status=400,
                )

        session.status = new_status
        session.save(update_fields=["status"])
        return Response(
            ExamSessionSerializer(session, context={"request": request}).data
        )


# ── Session Export ────────────────────────────────────────────────────────────


class SessionExportView(APIView):
    """
    GET /api/exam-sessions/export/?format=csv|xlsx&template=true
    """

    permission_classes = (IsAuthenticated,)

    HEADERS = [
        "course_code",
        "programme_code",
        "level_name",
        "date",
        "start_time",
        "end_time",
        "venue",
        "expected_students",
        "status",
    ]

    def get(self, request):
        fmt = request.query_params.get("format", "xlsx").lower()
        template_only = request.query_params.get("template") == "true"
        qs = (
            []
            if template_only
            else ExamSession.objects.select_related(
                "course", "programme", "level"
            ).all()
        )

        if fmt == "csv":
            return self._csv(qs, template_only)
        return self._xlsx(qs, template_only)

    def _row(self, s):
        return [
            s.course.course_code,
            s.programme.code,
            s.level.name,
            str(s.date),
            str(s.start_time)[:5],
            str(s.end_time)[:5] if s.end_time else "",
            s.venue,
            s.expected_students,
            s.status,
        ]

    def _xlsx(self, qs, template_only):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Sessions"
        ws.append(self.HEADERS)
        _style_header_row(ws, 1, len(self.HEADERS))
        for col, width in zip("ABCDEFGHI", [14, 16, 14, 12, 10, 10, 20, 18, 12]):
            ws.column_dimensions[col].width = width
        if not template_only:
            for s in qs:
                ws.append(self._row(s))
        name = "sessions_template.xlsx" if template_only else "exam_sessions.xlsx"
        return _xlsx_response(wb, name)


# ── Session Import ────────────────────────────────────────────────────────────


class SessionImportView(APIView):
    """
    POST /api/exam-sessions/import/
    Columns: course_code, programme_code, level_name, date (YYYY-MM-DD),
             start_time (HH:MM), end_time (HH:MM, optional),
             venue (optional), expected_students (optional)
    Sessions are created only (not upserted) to avoid overwriting live sessions.
    """

    permission_classes = (IsAdmin,)
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            rows = _parse_uploaded_file(file)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        if not rows:
            return Response({"detail": "The file is empty."}, status=400)

        courses = {c.course_code.upper(): c for c in Course.objects.all()}
        programmes = {p.code.upper(): p for p in Program.objects.all()}
        levels = {level.name.upper(): level for level in Level.objects.all()}

        created, errors = 0, []

        for i, row in enumerate(rows, start=1):
            course_code = row.get("course_code", "").strip().upper()
            programme_code = row.get("programme_code", "").strip().upper()
            level_name = row.get("level_name", "").strip().upper()
            date_str = row.get("date", "").strip()
            start_time_str = row.get("start_time", "").strip()
            end_time_str = row.get("end_time", "").strip()
            venue = row.get("venue", "").strip()
            expected = row.get("expected_students", "0").strip()

            missing = [
                f
                for f, v in [
                    ("course_code", course_code),
                    ("programme_code", programme_code),
                    ("level_name", level_name),
                    ("date", date_str),
                    ("start_time", start_time_str),
                ]
                if not v
            ]
            if missing:
                errors.append(
                    {
                        "row": i,
                        "error": f"Missing required fields: {', '.join(missing)}",
                    }
                )
                continue

            if course_code not in courses:
                errors.append({"row": i, "error": f"Course '{course_code}' not found."})
                continue
            if programme_code not in programmes:
                errors.append(
                    {"row": i, "error": f"Programme '{programme_code}' not found."}
                )
                continue
            if level_name not in levels:
                errors.append({"row": i, "error": f"Level '{level_name}' not found."})
                continue

            try:
                ExamSession.objects.create(
                    course=courses[course_code],
                    programme=programmes[programme_code],
                    level=levels[level_name],
                    date=date_str,
                    start_time=start_time_str,
                    end_time=end_time_str or None,
                    venue=venue,
                    expected_students=int(expected) if expected.isdigit() else 0,
                    status="scheduled",
                    created_by=request.user,
                )
                created += 1
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        return Response(
            {"created": created, "errors": errors},
            status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_200_OK,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Staff Users — Export / Import
# ─────────────────────────────────────────────────────────────────────────────


class StaffExportView(APIView):
    """
    GET /api/staff/export/?format=csv|xlsx&template=true
    """

    permission_classes = (IsAdmin,)

    HEADERS = ["username", "role", "phone_number", "password"]

    def get(self, request):
        fmt = request.query_params.get("format", "xlsx").lower()
        template_only = request.query_params.get("template") == "true"
        # Never export password hashes; template has password column for imports
        qs = [] if template_only else User.objects.all().order_by("username")

        if fmt == "csv":
            return self._csv(qs, template_only)
        return self._xlsx(qs, template_only)

    def _csv(self, qs, template_only):
        resp = HttpResponse(content_type="text/csv")
        name = "staff_template.csv" if template_only else "staff_users.csv"
        resp["Content-Disposition"] = f'attachment; filename="{name}"'
        w = csv.writer(resp)
        export_headers = (
            ["username", "role", "phone_number", "is_active", "created_at"]
            if not template_only
            else self.HEADERS
        )
        w.writerow(export_headers)
        if not template_only:
            for u in qs:
                w.writerow(
                    [
                        u.username,
                        getattr(u, "role", ""),
                        getattr(u, "phone_number", ""),
                        u.is_active,
                        u.date_joined.strftime("%Y-%m-%d")
                        if hasattr(u, "date_joined")
                        else "",
                    ]
                )
        return resp

    def _xlsx(self, qs, template_only):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Users"
        export_headers = (
            ["username", "role", "phone_number", "is_active", "created_at"]
            if not template_only
            else self.HEADERS
        )
        ws.append(export_headers)
        _style_header_row(ws, 1, len(export_headers))
        for col, width in zip("ABCDE", [20, 15, 18, 10, 14]):
            ws.column_dimensions[col].width = width
        if not template_only:
            for u in qs:
                ws.append(
                    [
                        u.username,
                        getattr(u, "role", ""),
                        getattr(u, "phone_number", ""),
                        u.is_active,
                        u.date_joined.strftime("%Y-%m-%d")
                        if hasattr(u, "date_joined")
                        else "",
                    ]
                )
        name = "staff_template.xlsx" if template_only else "staff_users.xlsx"
        return _xlsx_response(wb, name)


class StaffImportView(APIView):
    """
    POST /api/staff/import/
    Columns: username, role, phone_number (optional), password
    Creates new accounts only; existing usernames are skipped.
    """

    permission_classes = (IsAdmin,)
    parser_classes = [MultiPartParser]

    VALID_ROLES = {"admin", "invigilator"}

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            rows = _parse_uploaded_file(file)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        if not rows:
            return Response({"detail": "The file is empty."}, status=400)

        existing_usernames = set(User.objects.values_list("username", flat=True))
        created, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=1):
            username = row.get("username", "").strip()
            role = row.get("role", "invigilator").strip().lower()
            phone_number = row.get("phone_number", "").strip()
            password = row.get("password", "").strip()

            if not username:
                errors.append({"row": i, "error": "username is required."})
                continue
            if not password:
                errors.append({"row": i, "error": "password is required."})
                continue
            if role not in self.VALID_ROLES:
                errors.append(
                    {"row": i, "error": f"role must be one of {self.VALID_ROLES}."}
                )
                continue
            if username in existing_usernames:
                skipped += 1
                continue

            try:
                user = User.objects.create_user(username=username, password=password)
                if hasattr(user, "role"):
                    user.role = role
                if hasattr(user, "phone_number") and phone_number:
                    user.phone_number = phone_number
                user.save()
                existing_usernames.add(username)
                created += 1
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        return Response(
            {"created": created, "skipped": skipped, "errors": errors},
            status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_200_OK,
        )


# ─────────────────────────────────────────────────────────────────────────────
# QR Scan
# ─────────────────────────────────────────────────────────────────────────────


class ScanView(APIView):
    permission_classes = (CanScan,)

    def post(self, request):
        serializer = ScanSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        result = serializer.save()

        http_status = (
            status.HTTP_200_OK
            if result["status"] == "duplicate"
            else status.HTTP_201_CREATED
        )
        return Response(result, status=http_status)


# ─────────────────────────────────────────────────────────────────────────────
# Attendance Register
# ─────────────────────────────────────────────────────────────────────────────


class AttendanceListView(generics.ListAPIView):
    """GET /api/exam-sessions/<session_pk>/attendance/"""

    serializer_class = AttendanceSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter)
    filterset_fields = ("section", "status")
    search_fields = ("student__index_number", "student__full_name")

    def get_queryset(self):
        return (
            ExamAttendance.objects.filter(exam_session_id=self.kwargs["session_pk"])
            .select_related(
                "student", "student__programme", "student__level", "scanned_by"
            )
            .order_by("section", "scan_time")
        )


# ─────────────────────────────────────────────────────────────────────────────
# Attendance Export  (per-session — CSV or XLSX)
# ─────────────────────────────────────────────────────────────────────────────


class AttendanceExportView(APIView):
    permission_classes = (IsAuthenticated,)

    HEADERS = [
        "Section",
        "Index Number",
        "Full Name",
        "Programme",
        "Level",
        # "Gender",
        # "Scan Time",
        # "Status",
        # "Scanned By",
    ]

    def get(self, request, session_pk):

        session = get_object_or_404(
            ExamSession.objects.select_related("course", "programme", "level"),
            pk=session_pk,
        )

        section_filter = request.query_params.get("section")

        qs = (
            ExamAttendance.objects.filter(exam_session=session)
            .select_related(
                "student", "student__programme", "student__level", "scanned_by"
            )
            .order_by("section", "scan_time")
        )

        if section_filter in ("A", "B"):
            qs = qs.filter(section=section_filter)

        return self._pdf(session, qs)

    def _record_row(self, record):
        s = record.student
        return [
            record.get_section_display(),
            s.index_number,
            s.full_name,
            s.programme.name,
            s.level.name,
            # s.gender,
            # record.scan_time.strftime("%Y-%m-%d %H:%M:%S"),
            # record.get_status_display(),
            # record.scanned_by.username,
        ]

    def _pdf(self, session, qs):

        filename = f"attendance_{session.course.course_code}_{session.date}.pdf"

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        styles = getSampleStyleSheet()
        subtitle_style = ParagraphStyle(
            "CenteredSubtitle", parent=styles["Normal"], alignment=TA_CENTER
        )

        elements = []

        # Title
        elements.append(
            Paragraph("<b>EXAMINATION ATTENDANCE REGISTER</b>", styles["Title"])
        )

        elements.append(Spacer(1, 10))

        # Session metadata
        elements.append(
            Paragraph(
                f"{session.course.course_code} – {session.course.course_title}<br/>"
                f"{session.programme.name} | Level {session.level.name}",
                subtitle_style,
            )
        )

        end_time = session.end_time.strftime("%H:%M") if session.end_time else "—"

        elements.append(
            Paragraph(
                f"Date: {session.date} | Venue: {session.venue or '—'} | "
                f"{session.start_time.strftime('%H:%M')} – {end_time}",
                subtitle_style,
            )
        )

        elements.append(Spacer(1, 20))

        # Table data
        table_data = [self.HEADERS]

        for record in qs:
            table_data.append(self._record_row(record))

        table = Table(table_data, repeatRows=1)

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (1, 1), (1, -1), "CENTER"),
                ]
            )
        )

        elements.append(table)

        doc.build(elements)

        return response


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────────────────────


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard(request):
    today = timezone.now().date()
    data = {
        "total_students": Student.objects.filter(is_active=True).count(),
        "total_sessions": ExamSession.objects.count(),
        "active_sessions": ExamSession.objects.filter(status="active").count(),
        "total_scans_today": ExamAttendance.objects.filter(
            scan_time__date=today
        ).count(),
    }
    return Response(data)
